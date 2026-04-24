import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
 
#Lectura 
INPUT_FILE  = r"C:\Users\tamar\Documents\Capstone\Documentos Generados\ChocloNorteTotal.xlsx"
OUTPUT_FILE = r"C:\Users\tamar\Documents\Capstone\Documentos Generados\ChocloNorte_Limpio.xlsx"
SHEET_DETALLE = "Datos Producto"
 
df_raw = pd.read_excel(INPUT_FILE, sheet_name=SHEET_DETALLE)
df_ts  = pd.read_excel(INPUT_FILE, sheet_name="Serie Temporal")
 
#Limpieza
#convertir negativos → 0 (son devoluciones)
mask_neg = df_raw["Demanda"] < 0
df_clean = df_raw.copy()
df_clean.loc[mask_neg, "Demanda"] = 0
 
#eliminar registros con Demanda == 0
mask_zero = df_clean["Demanda"] == 0
df_clean = df_clean[~mask_zero].reset_index(drop=True)
 
 
#Serie temporal reconstruida 
ts_nueva = (
    df_clean
    .groupby("Mes", as_index=False)["Demanda"]
    .sum()
    .rename(columns={"Demanda": "Demanda Limpia"})
)
ts_comparacion = df_ts.copy()
ts_comparacion.columns = ["Mes", "Demanda Original"]
ts_comparacion = ts_comparacion.merge(ts_nueva, on="Mes", how="left")
ts_comparacion["Diferencia"] = (
    ts_comparacion["Demanda Limpia"] - ts_comparacion["Demanda Original"]
)
 
#Guardar Archivo en formato xlsx
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df_clean.to_excel(writer, sheet_name="Datos Limpios",       index=False)
    ts_comparacion.to_excel(writer, sheet_name="Serie Temporal", index=False)
    # log.to_excel(writer,       sheet_name="Log de Cambios",     index=False)
 
#Resumen en consola
print("=" * 55)
print("           RESUMEN DE LIMPIEZA DE DATOS")
print("=" * 55)
print(f"  Registros originales       : {len(df_raw):>6,}")
print(f"  Negativos convertidos a 0  : {mask_neg.sum():>6,}")
print(f"  Ceros eliminados (total)   : {(df_raw['Demanda'] == 0).sum() + mask_neg.sum():>6,}")
print(f"  Registros finales limpios  : {len(df_clean):>6,}")
print(f"  Archivo generado           : {OUTPUT_FILE}")
print("=" * 55)