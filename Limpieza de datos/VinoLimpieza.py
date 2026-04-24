import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
 
# ── 1. Lectura ────────────────────────────────────────────────────────────────
INPUT_FILE  = r"C:\Users\tamar\Documents\Capstone\Documentos Generados\VinoTotal.xlsx"
OUTPUT_FILE = r"C:\Users\tamar\Documents\Capstone\Documentos Generados\Vino_Limpio.xlsx"
SHEET_DETALLE = "Datos Producto"
 
df_raw = pd.read_excel(INPUT_FILE, sheet_name=SHEET_DETALLE)
df_ts  = pd.read_excel(INPUT_FILE, sheet_name="Serie Temporal")
 
# ── 2. Limpieza ───────────────────────────────────────────────────────────────
# Paso A: convertir negativos → 0 (son devoluciones)
mask_neg = df_raw["Demanda"] < 0
df_clean = df_raw.copy()
df_clean.loc[mask_neg, "Demanda"] = 0
 
# Paso B: eliminar registros con Demanda == 0
mask_zero = df_clean["Demanda"] == 0
df_clean = df_clean[~mask_zero].reset_index(drop=True)
 
# ── 3. Log de cambios ─────────────────────────────────────────────────────────
# log_neg  = df_raw[mask_neg][["NomClienteAlter","Mes","Demanda"]].copy()
# log_neg.columns = ["Cliente","Mes","Demanda Original"]
# log_neg["Accion"] = "Negativo → 0 → eliminado"
 
# log_zero = df_raw[(df_raw["Demanda"] == 0)][["NomClienteAlter","Mes","Demanda"]].copy()
# log_zero.columns = ["Cliente","Mes","Demanda Original"]
# log_zero["Accion"] = "Cero → eliminado"
 
# log = pd.concat([log_neg, log_zero], ignore_index=True).sort_values("Mes")
 
# ── 4. Serie temporal reconstruida ────────────────────────────────────────────
ts_nueva = (
    df_clean
    .groupby("Mes", as_index=False)["Demanda"]
    .sum()
    .rename(columns={"Demanda": "Demanda Limpia"})
)
ts_comparacion = df_ts.copy()
ts_comparacion.columns = ["Mes", "Demanda Original"]
ts_comparacion = ts_comparacion.merge(ts_nueva, on="Mes", how="left")
ts_comparacion["Diferencia"] = (
    ts_comparacion["Demanda Limpia"] - ts_comparacion["Demanda Original"]
)
 
# ── 5. Escritura con openpyxl ─────────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df_clean.to_excel(writer, sheet_name="Datos Limpios",       index=False)
    ts_comparacion.to_excel(writer, sheet_name="Serie Temporal", index=False)
    # log.to_excel(writer,       sheet_name="Log de Cambios",     index=False)
 
# ── 6. Formato ────────────────────────────────────────────────────────────────
# wb = load_workbook(OUTPUT_FILE)
 
# HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
# HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
# CHANGED_FILL  = PatternFill("solid", fgColor="FFE699")   # amarillo suave
# NEG_FILL      = PatternFill("solid", fgColor="F4CCCC")   # rojo suave
# ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")   # azul muy claro
 
# def format_sheet(ws, col_widths=None):
#     for cell in ws[1]:
#         cell.font      = HEADER_FONT
#         cell.fill      = HEADER_FILL
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#     ws.row_dimensions[1].height = 20
#     for row in ws.iter_rows(min_row=2):
#         for cell in row:
#             cell.font      = Font(name="Arial", size=10)
#             cell.alignment = Alignment(horizontal="left")
#     if col_widths:
#         for col_letter, width in col_widths.items():
#             ws.column_dimensions[col_letter].width = width
 
# # Hoja Datos Limpios
# ws_d = wb["Datos Limpios"]
# format_sheet(ws_d, {"A":42,"B":14,"C":26,"D":14,"E":12})
# for i, row in enumerate(ws_d.iter_rows(min_row=2), start=2):
#     fill = ALT_FILL if i % 2 == 0 else None
#     for cell in row:
#         if fill:
#             cell.fill = fill
 
# # Hoja Serie Temporal
# ws_t = wb["Serie Temporal"]
# format_sheet(ws_t, {"A":20,"B":18,"C":18,"D":14})
# diff_col = 4  # columna D = Diferencia
# for row in ws_t.iter_rows(min_row=2):
#     diff_cell = row[diff_col - 1]
#     if diff_cell.value and diff_cell.value != 0:
#         for cell in row:
#             cell.fill = CHANGED_FILL
 
# # Hoja Log de Cambios
# ws_l = wb["Log de Cambios"]
# format_sheet(ws_l, {"A":42,"B":14,"C":18,"D":28})
# for row in ws_l.iter_rows(min_row=2):
#     orig = row[2].value
#     fill = NEG_FILL if (orig is not None and orig < 0) else CHANGED_FILL
#     for cell in row:
#         cell.fill = fill
 
# wb.save(OUTPUT_FILE)
 
# ── 7. Resumen en consola ─────────────────────────────────────────────────────
print("=" * 55)
print("           RESUMEN DE LIMPIEZA DE DATOS")
print("=" * 55)
print(f"  Registros originales       : {len(df_raw):>6,}")
print(f"  Negativos convertidos a 0  : {mask_neg.sum():>6,}")
print(f"  Ceros eliminados (total)   : {(df_raw['Demanda'] == 0).sum() + mask_neg.sum():>6,}")
print(f"  Registros finales limpios  : {len(df_clean):>6,}")
print(f"  Archivo generado           : {OUTPUT_FILE}")
print("=" * 55)